# create_fixed_template.py
# Excel inspection template with:
# - Title: no border
# - Notes: underlined, no border, top-left
# - BIN / Team Leader / Span / Inspection Date / Weather / Asst. Team Leader: underlined value cells
# - Condition block now has real VALUE areas for Note, Condition State, References Photo(s), References Sketch(es)
#   so your converter has places to write (with matching anchors in _anchors)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Styles ----------
MEDIUM = Side(style="medium", color="000000")
THIN   = Side(style="thin",   color="000000")
OUTLINE = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_MID = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ---------- Helpers ----------
def outline(ws, r1, c1, r2, c2, side=MEDIUM):
    """Draw only the OUTER rectangle border around r1..r2, c1..c2 (1-based)."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            left   = side if c == c1 else cell.border.left
            right  = side if c == c2 else cell.border.right
            top    = side if r == r1 else cell.border.top
            bottom = side if r == r2 else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def merge_label(ws, cell_range, text, font=None, align=None, bold=False, underline=None):
    """Merge a range and place a single formatted label in the top-left cell."""
    ws.merge_cells(cell_range)
    tl = cell_range.split(":")[0]
    c = ws[tl]
    c.value = text
    if font is None:
        font = Font()
    if bold or underline:
        font = Font(bold=bold, underline=underline)
    c.font = font
    if align:
        c.alignment = align
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def underline_bottom(ws, cell_ref, side=THIN):
    """Add a visible underline (bottom border) to a single cell."""
    cell = ws[cell_ref]
    b = cell.border
    cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=side)

# ---------- Template ----------
def build_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "TEMPLATE"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [6,14,14,14,14,14,14,14,14,14,14,14,14])

    # -------- Title (NO BORDER) --------
    merge_label(
        ws, "B2:L2",
        "2022 GOWANUS EXPRESSWAY IN-DEPTH INSPECTION",
        font=Font(size=16, bold=True),
        align=CENTER
    )

    # -------- Header block (Rows 4–7) --------
    outline(ws, 4, 2, 7, 13)
    ws["B4"].value = "BIN:"; ws["L4"].value = "Inspection Date:"
    ws["L6"].value = "Weather:"; ws["B6"].value = "Team Leader:"
    ws["L7"].value = "Asst. Team Leader:"; ws["B7"].value = "Span:"
    ws["D7"].value = "Location:"
    for addr in ["B4","L4","L6","B6","L7","B7","D7"]:
        ws[addr].font = Font(bold=True)
        ws[addr].alignment = LEFT_MID

    # Underlined value cells (top-left cells of the write-in areas)
    for addr in ["C4", "C6", "C7", "M4", "M6", "M7", "E7"]:
        underline_bottom(ws, addr, THIN)
        ws[addr].alignment = LEFT_MID

    # -------- Notes (NO BORDER, underlined label) --------
    merge_label(ws, "B10:L11", "Notes:", bold=True, underline="single", align=LEFT_TOP)

    # -------- Condition block (Rows 13–18) --------
    # Create real VALUE areas at C..G for each label so your script can write
    outline(ws, 13, 2, 18, 13)

    # left labels + value ranges
    ws["B13"] = "Location:";            ws["B13"].font = Font(bold=True); ws["B13"].alignment = LEFT_MID
    ws["B15"] = "Note:";                ws["B15"].font = Font(bold=True); ws["B15"].alignment = LEFT_MID
    ws["B16"] = "Condition State:";     ws["B16"].font = Font(bold=True); ws["B16"].alignment = LEFT_MID
    ws["B17"] = "References Photo(s):"; ws["B17"].font = Font(bold=True); ws["B17"].alignment = LEFT_MID
    ws["B18"] = "References Sketch(es)";ws["B18"].font = Font(bold=True); ws["B18"].alignment = LEFT_MID

    ws["B9"] = "Notes:"
    ws["B9"].font = Font(bold=True, underline="single")
    ws["B9"].alignment = LEFT_TOP
    ws.merge_cells("B10:L11")


    # value areas (merged C..G)
    ws.merge_cells("C13:G13")
    ws.merge_cells("C16:G16")
    ws.merge_cells("C17:G17")
    ws.merge_cells("C18:G18")

    # CS0..CS5 headers & mark row (H..M)
    for col, label in zip("HIJKLM", ["CS0","CS1","CS2","CS3","CS4","CS5"]):
        c = ws[f"{col}13"]
        c.value = label
        c.alignment = CENTER
        c.font = Font(bold=True)
        c.border = OUTLINE
    for col in "HIJKLM":
        ws[f"{col}14"].alignment = CENTER
        ws[f"{col}14"].border = OUTLINE

    # -------- Description (NO BORDER) --------
    # Label on its own cell (no merge)
    ws["B19"] = "Description:"
    ws["B19"].font = Font(bold=True)
    ws["B19"].alignment = LEFT_TOP

    # Value area merged separately (starts at B20)
    ws.merge_cells("B20:M22")

    # -------- Photos Section --------
    merge_label(ws, "B23:M23", "Inspection Photographs", bold=True, align=CENTER)
    outline(ws, 23, 2, 23, 13)
    ws.row_dimensions[23].height = 22

    # Left: Attachment Description box
    merge_label(ws, "B26:D38", "Attachment Description:", bold=True, align=LEFT_MID)
    outline(ws, 25, 2, 38, 4)

    # Row 25 meta labels (right)
    # ---- Photos meta row (Row 25) ----
    # Labels in single, unmerged cells
    ws["G25"] = "Photo Number:"
    ws["G25"].font = Font(bold=True);
    ws["G25"].alignment = LEFT_MID

    ws["J25"] = "Photo Filename:"
    ws["J25"].font = Font(bold=True);
    ws["J25"].alignment = LEFT_MID

    # Value areas (own merges, top-left is the anchor)
    ws.merge_cells("H25:I25")  # value area for Photo Number   -> anchor H25
    ws.merge_cells("K25:L25")  # value area for Photo Filename -> anchor J25

    # Keep the band outline
    outline(ws, 25, 2, 26, 13)

    # Photo rectangle (E27:M38)
    outline(ws, 27, 5, 38, 13)

    # ---------- Hidden anchors ----------
    # Add anchors for the new value areas we created in C..G so your converter can write!
    anchors = {
        # Header block
        "BIN":"C4","Inspection Date":"M4","Weather":"M6",
        "Team Leader":"C6","Asst Team Leader":"M7","Span":"C7","Location":"E7",
        # Notes + Condition block
        "Notes":"B10",
        "Condition Location":"C13",
        "Condition Note":"C15",
        "Condition State:":"C16",             # note the colon matches converter
        "References Photo(s):":"C17",
        "References Sketch(es)":"C18",
        # CS buckets
        "CS0":"H14","CS1":"I14","CS2":"J14","CS3":"K14","CS4":"L14","CS5":"M14",
        # Narrative / Photos area
        "Description":"B20",
        "Attachment Description":"B26",
        "Photo Number":"H25",
        "Photo Filename":"K25",
    }

    meta = wb.create_sheet("_anchors")
    meta["A1"] = "field"; meta["B1"] = "cell"
    for i, (k, v) in enumerate(anchors.items(), start=2):
        meta[f"A{i}"] = k; meta[f"B{i}"] = v
    meta.sheet_state = "hidden"

    return wb

if __name__ == "__main__":
    wb = build_template()
    wb.save("inspection_template2.xlsx")
    print("Wrote inspection_template2.xlsx")

