# excel_to_report_wsp.py
# Reads: inputexcelfile.xlsx
# Template: inspection_template2.xlsx
# Output: inspection_reports.xlsx

import os, re, time
from pathlib import Path 
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage

# Sizing (Excel/openpyxl uses pixels; assume 96 dpi)
DPI = 96
PHOTO_W_IN = 3.0     # width in inches
PHOTO_H_IN = 3.5     # height in inches
PHOTO_W_PX = int(PHOTO_W_IN * DPI)  # 288 px
PHOTO_H_PX = int(PHOTO_H_IN * DPI)  # 336 px


# ---- Template keys must match hidden _anchors in template ----
TEMPLATE_KEYS = [
    "BIN","Inspection Date","Team Leader","Asst Team Leader","Span","Location","Weather",
    "Notes",
    "Condition Location","Condition Note","Condition State:",
    "References Photo(s):","References Sketch(es)",
    "CS0","CS1","CS2","CS3","CS4","CS5",
    "Description","Attachment Description","Photo Number","Photo Filename","Photo Path"
]

def _canon(s): return re.sub(r'[^a-z0-9]', '', str(s or '').strip().lower())

VARIANT_MAP = {
    "bin":"BIN","inspectiondate":"Inspection Date",
    "teamleader":"Team Leader","asstteamleader":"Asst Team Leader",
    "assistantteamleader":"Asst Team Leader","span":"Span",
    "location":"Location","weather":"Weather","notes":"Notes",
    "conditionlocation":"Condition Location","member":"Condition Location",
    "conditionnote":"Condition Note","conditionstate":"Condition State:",
    "conditionstate:":"Condition State:","condition":"Condition State:",
    "referencesphotos":"References Photo(s):","referencesphoto(s)":"References Photo(s):",
    "referencesketches":"References Sketch(es)","referencesketch(es)":"References Sketch(es)",
    "cs0":"CS0","cs1":"CS1","cs2":"CS2","cs3":"CS3","cs4":"CS4","cs5":"CS5",
    "narrative":"Description","description":"Description",
    "attachmentdescription":"Attachment Description","photonumber":"Photo Number",
    "photofilename":"Photo Filename","photopath":"Photo Path"
}

def _normalize_dataframe(df):
    rename = {}
    for col in list(df.columns):
        key = VARIANT_MAP.get(_canon(col))
        if key: rename[col] = key
    df2 = df.rename(columns=rename).copy()
    for k in TEMPLATE_KEYS:
        if k not in df2.columns: df2[k] = ""
    return df2

def _safe_title(name): return re.sub(r'[\[\]\:\*\?\/\\]', '_', str(name))[:31]

def _read_anchors(wb):
    ws = wb["_anchors"]
    return {str(r[0]): str(r[1]) for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0] and r[1]}

def _coerce_txt(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return ""
    if isinstance(v, float) and float(v).is_integer(): return str(int(v))
    return str(v)

IMAGE_EXTS = [".jpg",".jpeg",".png",".bmp",".tif",".tiff",".gif",
              ".JPG",".JPEG",".PNG",".BMP",".TIF",".TIFF",".GIF"]

def _split_list(s):
    if s is None or (isinstance(s, float) and pd.isna(s)): return []
    s = str(s).strip()
    if not s: return []
    return [x.strip().strip('"').strip("'") for x in re.split(r"[;,|]", s)]

def resolve_photo_files(photo_path_value, photo_filename_value, base_dir: Path):
    """
    photo_path_value  : folder path (e.g., C:\\Users\\Hp\\Desktop\\excel_to_excel\\photo)
    photo_filename_value : name(s) like 'AA_113_2933' or 'photo1.jpg;photo2.jpg'
    Returns a list of existing image file paths.
    """
    folder = str(photo_path_value or "").strip()
    names  = _split_list(photo_filename_value)
    if not folder or not names:
        return []

    folder = Path(folder)
    if not folder.is_absolute():
        folder = (base_dir / folder).resolve()

    out = []
    for name in names:
        p = folder / name
        # exact name (with extension) given?
        if p.exists() and p.is_file():
            out.append(str(p)); continue
        # try appending common extensions if none given
        if not Path(name).suffix:
            found = False
            for ext in IMAGE_EXTS:
                p2 = folder / f"{name}{ext}"
                if p2.exists() and p2.is_file():
                    out.append(str(p2)); found = True; break
            if found: continue
        # try prefix match (e.g., AA_113_2933*.jpg)
        for ext in IMAGE_EXTS:
            for cand in folder.glob(f"{name}*{ext.lower()}"):
                if cand.is_file(): out.append(str(cand))
            for cand in folder.glob(f"{name}*{ext.upper()}"):
                if cand.is_file(): out.append(str(cand))
    # dedupe preserve order; limit to 2 since you want E27 and M27
    seen, final = set(), []
    for p in out:
        if p not in seen:
            final.append(p); seen.add(p)
        if len(final) >= 2: break
    return final

# target size
DPI = 96
W_PX = int(3.0 * DPI)   # width  3.0"
H_PX = int(3.5 * DPI)   # height 3.5"

def place_images_at_E27_M27(ws, image_paths):
    """Place up to 2 images: first at E27, second at M27. Fit to 3.0\"x3.5\"."""
    anchors = ["E27", "M27"]  # exact anchors you requested
    for pos, p in zip(anchors, image_paths):
        try:
            img = XLImage(p)
            ow, oh = int(img.width), int(img.height)
            if ow and oh:
                s = min(W_PX/ow, H_PX/oh)
                img.width  = max(1, int(ow*s))
                img.height = max(1, int(oh*s))
            ws.add_image(img, pos)
        except Exception:
            # ignore a bad image and continue
            continue

def build_output(input_excel, template_excel, output_excel):
    input_excel, template_excel, output_excel = map(str, (input_excel, template_excel, output_excel))
    df = pd.read_excel(input_excel, dtype=object)
    if df.empty: raise ValueError("No data in input file.")
    df = _normalize_dataframe(df)
    wb = load_workbook(template_excel)
    anchors = _read_anchors(wb)
    t_ws = wb["TEMPLATE"]
    base_dir = Path(input_excel).resolve().parent

    for i, rec in enumerate(df.to_dict(orient="records"), start=1):
        ws = wb.copy_worksheet(t_ws)
        ws.sheet_view.showGridLines = False
        title = _safe_title(f"{rec.get('BIN','')}_{i}") or f"Report_{i}"
        ws.title = title
        # Fill all fields
        for k in TEMPLATE_KEYS:
            if k=="Photo Path": continue
            if k in anchors:
                c = anchors[k]
                ws[c].value = _coerce_txt(rec.get(k,""))
                ws[c].alignment = Alignment(wrap_text=True, vertical="top")
        base_dir = Path(input_excel).resolve().parent

        photo_files = resolve_photo_files(
            rec.get("Photo Path"),  # e.g., C:\Users\Hp\Desktop\excel_to_excel\photo
            rec.get("Photo Filename"),  # e.g., AA_113_2933  (or 'photo1.jpg;photo2.jpg')
            base_dir
        )
        place_images_at_E27_M27(ws, photo_files)

    # Remove template and anchors
    for s in ["TEMPLATE","_anchors"]:
        if s in wb.sheetnames: wb.remove(wb[s])
    try:
        wb.save(output_excel)
    except PermissionError:
        ts = int(time.time())
        wb.save(f"{Path(output_excel).stem}_{ts}.xlsx")
    print(f" Report generated: {output_excel}")

# ------------------------------
# MAIN ENTRY (no argparse)
# ------------------------------
if __name__ == "__main__":
    base = Path(__file__).resolve().parent
    print("Working directory:", base)
    input_x  = base/"inputexcelfile.xlsx"
    templ_x  = base/"inspection_template2.xlsx"
    output_x = base/"inspection_reports.xlsx"

    build_output(input_x, templ_x, output_x)


